import configparser
import logging
import platform
import json
import sys

from os import sep, path, remove

from openpyxl import Workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.common.action_chains import ActionChains

import bs4
import urllib3

from urllib.parse import urlsplit

http = urllib3.PoolManager()
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

dir_path = path.dirname(path.realpath(__file__))
sys.path.append(dir_path + "modules")
sys.path.append(dir_path + "drivers")


class Configs:
    file = r"./configs.txt"

    config = {"threads": 1, "browser": "chrome"}
    parsed = False

    @staticmethod
    def parse_config_file():
        config_parser = configparser.RawConfigParser()
        config_parser.read(Configs.file)

        # scraper configs
        Configs.config['driver'] = config_parser.get('scraper', 'driver')
        Configs.config['html_parser'] = config_parser.get('scraper', 'html_parser')
        Configs.config['logging_handler'] = config_parser.get('scraper', 'logging_handler')
        Configs.config['threads'] = config_parser.getint('scraper', 'threads')
        Configs.config['output_format'] = config_parser.get('scraper', 'output_format')
        Configs.config['testing'] = config_parser.getboolean('scraper', 'testing')
        Configs.config['max_browsers'] = config_parser.getint('scraper', 'browsers')
        Configs.config['max_items_extract'] = config_parser.getint('scraper', 'max_items_extract')
        Configs.config['website_url'] = config_parser.get('scraper', 'website_url')

        # pagination
        Configs.config['use_selenium'] = config_parser.getboolean('pagination', 'use_selenium')
        Configs.config['wait_before_pagination'] = config_parser.getint('pagination', 'wait_before_pagination')
        Configs.config['wait_after_pagination'] = config_parser.getint('pagination', 'wait_after_pagination')
        Configs.config['pagination_tag'] = config_parser.get('pagination', 'pagination_tag')
        Configs.config['pagination_attribute'] = config_parser.get('pagination', 'pagination_attribute')
        Configs.config['pagination_attribute_value'] = config_parser.get('pagination', 'pagination_attribute_value')
        Configs.config['pagination_xpath'] = config_parser.get('pagination', 'pagination_xpath')

        # listings
        Configs.config['listing_tag'] = config_parser.get('listings', 'listing_tag')
        Configs.config['listing_attribute'] = config_parser.get('listings', 'listing_attribute')
        Configs.config['listing_attribute_value'] = config_parser.get('listings', 'listing_attribute_value')

        # profile items
        Configs.config['item_ids'] = dict(config_parser.items('profile items ids'))
        Configs.config['item_xpaths'] = dict(config_parser.items('profile items xpaths'))

        Configs.read = True

    @staticmethod
    def get(key, check_for_none=True):
        if not Configs.parsed:
            Configs.parse_config_file()

        cfg = Configs.config[key]

        if check_for_none and cfg is None:
            logging.error('''Please specify \'{}\' in configs.txt file!'''.format(key))
            exit(1)

        return cfg


def get_domain(url):
    domain = "{0.scheme}://{0.netloc}/".format(urlsplit(url))
    assert domain
    return domain


# end of configs class
def configure_logging():
    rootLogger = logging.getLogger()
    logFormatter = logging.Formatter("%(filename)s:%(lineno)s %(asctime)s [%(levelname)-5.5s]  %(message)s")

    if "file" in str(Configs.get("logging_handler")):
        filename = dir_path + sep + "scraper.log"
        remove(filename) if path.exists(filename) else None
        handler = logging.FileHandler(filename=filename)
    else:
        handler = logging.StreamHandler()

    handler.setFormatter(logFormatter)
    rootLogger.addHandler(handler)
    rootLogger.setLevel(logging.INFO)


def setup_browser(browser=""):
    if browser == "":
        browser = Configs.get("driver")
    bpath = dir_path + sep + "drivers" + sep + browser

    if "Windows" in platform.system():
        bpath += ".exe"

    if "chrome" in browser:
        bpath = bpath.replace("chrome", "chromedriver")
        driver = setup_chrome(bpath)
    elif "phantomjs" in browser:
        driver = setup_phantomjs(bpath)
    elif "firefox" in browser:
        bpath = bpath.replace("firefox", "geckodriver")
        driver = setup_firefox(bpath)
    else:
        driver = setup_chrome(bpath)
        logging.warning("Invalid browser name specified, using default browser")

    return driver


def setup_chrome(bpath, maximize=True):
    opt = webdriver.ChromeOptions()

    opt.add_argument("--start-maximized")
    # disable images
    # disable_all = {#"profile.managed_default_content_settings.images": 2,
    # "profile.managed_default_content_settings.javascript": 2,
    # "profile.managed_default_content_settings.plugin": 2,
    # "profile.managed_default_content_settings.popups": 2,
    # "profile.managed_default_content_settings.automaticDownloads": 2}

    # opt.add_experimental_option("prefs", disable_all)
    # not sure that this work
    driver = webdriver.Chrome(bpath, chrome_options=opt)
    driver.delete_all_cookies()

    # maximize browser
    # if maximize:
    #   driver.maximize_window()
    return driver


def setup_firefox(bpath, maximize=True):
    firefox_profile = webdriver.FirefoxProfile()

    # disable css
    firefox_profile.set_preference('permissions.default.stylesheet', 2)
    # disable images
    firefox_profile.set_preference('permissions.default.image', 2)
    # disable flash
    firefox_profile.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')
    # disable javascript
    firefox_profile.set_preference('permissions.default.image', 2)

    driver = webdriver.Firefox(executable_path=bpath)  # , firefox_profile=firefox_profile)

    # maximize browser
    if maximize:
        driver.maximize_window()
    return driver


def setup_phantomjs(bpath, maximize=True):
    service_args = ['--ignore-ssl-errors=true', '--ssl-protocol=any']
    driver = webdriver.PhantomJS(bpath, service_args=service_args)
    if maximize:
        driver.maximize_window()
    return driver


def write_output(file_name, data):
    file_format = Configs.get("output_format")
    if file_format == "json":
        write_json_file(file_name.rsplit(".", 1)[0] + ".json", data)
    else:
        if file_format != "excel":
            logging.warning("Unknown output format is specified, using excel by default")

        write_to_excel(file_name.rsplit(".", 1)[0] + ".xlsx", data)


def write_json_file(name, data):
    with open(name, 'w') as fname:
        json.dump(data, fname)


def write_to_excel(xlsx_file, dict_list=None, sheet_title_1=None):
    if dict_list is None:
        dict_list = []
        logging.warning('Warning: No data was available for writing into the worksheet {}'.format(sheet_title_1))

    wb = Workbook(write_only=False)
    wb.guess_types = True
    ws = wb.create_sheet(title=sheet_title_1)
    del wb['Sheet']

    records = []
    for d in dict_list:
        records.append(list(d.values()))

    ws.append(list(dict_list[0].keys()))
    for record in records:
        ws.append(record)
    wb.save(xlsx_file)


def write_to_excel_reservation(xlsx_file, dict_list=None, sheet_title_1=None):
    if dict_list is None:
        dict_list = []
        logging.warning('Warning: No data was available for writing into the worksheet {}'.format(sheet_title_1))

    wb = Workbook(write_only=False)
    wb.guess_types = True
    ws = wb.create_sheet(title=sheet_title_1)
    del wb['Sheet']

    records = []
    for d in dict_list:
        pair = []
        for k, v in d.items():
            from_ = k.split("to")[0]
            to_ = k.split("to")[1]
            pair = [from_, to_, v]
        records.append(pair)

    ws.append(["Date from", "Date to", "Occupancy Ratio"])
    for record in records:
        ws.append(record)
    wb.save(xlsx_file)


def read_excel_file(excel_file):
    xls = ExcelFile(excel_file)
    return xls.parse(xls.sheet_names[0])


def append_into_file(file, string):
    with open(file, "a", encoding='utf-8') as myfile:
        myfile.write(string + '\n')


def write_lines_to_file(name, urls):
    with open(name, 'w', encoding='utf-8') as f:
        for url in urls:
            try:
                f.write(url + '\n')
            except Exception as e:
                print(str(e))


def load_page(url):
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1'}
    response = http.request('GET', url, headers=header)

    # return bs4.BeautifulSoup(response.data, 'html5lib')
    return bs4.BeautifulSoup(response.data, 'lxml')


# Clicks element
def click(driver, elem):
    try:
        elem.click()
        return True
    except WebDriverException:
        return False
    except:
        try:
            actions = ActionChains(driver)
            actions.move_to_element(elem)
            actions.click(elem)
            actions.perform()
        except WebDriverException:
            return False
